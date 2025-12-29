from flask import Blueprint, render_template, render_template_string, request, redirect, url_for, session
from openpyxl import load_workbook
import os

admin_bp = Blueprint("admin", __name__)

ADMIN_USER = "admin"
ADMIN_PASS = "1234"
TRAINER_FILE = "trainers.xlsx"

def load_trainers():
    if not os.path.exists(TRAINER_FILE):
        return [], []
    wb = load_workbook(TRAINER_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return rows[0], rows[1:]

@admin_bp.route("/admin", methods=["GET","POST"])
def admin_login():
    error = None
    if request.method == "POST":
        if request.form["username"] == ADMIN_USER and request.form["password"] == ADMIN_PASS:
            session["admin"] = True
            return redirect(url_for("admin.dashboard"))
        else:
            error = "Invalid login"
    return render_template_string("""
    <h2>Admin Login</h2>
    <form method="post">
      Username:<input name="username"><br><br>
      Password:<input type="password" name="password"><br><br>
      <button>Login</button>
    </form>
    <p style="color:red;">{{error}}</p>
    """, error=error)

@admin_bp.route("/admin/dashboard")
def dashboard():
    if not session.get("admin"):
        return redirect(url_for("admin.admin_login"))
    return render_template_string("""
    <h2>Admin Dashboard</h2>
    <ul>
      <li><a href="/trainers">View Trainers</a></li>
      <li><a href="/">Trainer Form</a></li>
      <li><a href="/logout">Logout</a></li>
    </ul>
    """)

@admin_bp.route("/trainers")
def view_trainers():
    if not session.get("admin"):
        return redirect(url_for("admin.admin_login"))

    headers, data = load_trainers()
    return render_template(
        "admin_trainers.html",
        headers=headers,
        data=data,
        enumerate=enumerate
    )
@admin_bp.route("/approve/<int:index>")
def approve(index):
    if not session.get("admin"):
        return redirect(url_for("admin.admin_login"))
    wb = load_workbook(TRAINER_FILE)
    ws = wb.active
    ws.cell(row=index+2, column=18, value="Approved")
    wb.save(TRAINER_FILE)
    return redirect(url_for("admin.view_trainers"))

@admin_bp.route("/logout")
def logout():
    session.pop("admin", None)
    return redirect(url_for("admin.admin_login"))
