import os, random, string
from openpyxl import load_workbook
from datetime import datetime

TRAINER_FILE = "trainers.xlsx"

def calculate_age(dob):
    dob = datetime.strptime(dob, "%Y-%m-%d")
    today = datetime.today()
    return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))

def generate_trainer_id():
    if not os.path.exists(TRAINER_FILE):
        return "TRN0001"

    wb = load_workbook(TRAINER_FILE)
    ws = wb.active
    last_row = ws.max_row

    if last_row < 2:
        return "TRN0001"

    last_id = ws.cell(row=last_row, column=17).value
    if last_id and str(last_id).startswith("TRN"):
        num = int(str(last_id).replace("TRN", ""))
        return f"TRN{num+1:04d}"
    return "TRN0001"

def generate_password(length=8):
    chars = string.ascii_letters + string.digits
    return ''.join(random.choice(chars) for _ in range(length))
