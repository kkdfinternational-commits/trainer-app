from flask import Blueprint, render_template, request, redirect, url_for
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
import os

from utils import calculate_age, generate_trainer_id, generate_password

forms_bp = Blueprint("forms", __name__)

UPLOAD_FOLDER = "uploads"
TRAINER_FILE = "trainers.xlsx"
SCHOOL_FILE = "schools.xlsx"

def get_used_schools():
    used = set()
    if os.path.exists(TRAINER_FILE):
        wb = load_workbook(TRAINER_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            schools_cell = row[11]
            if schools_cell:
                for s in str(schools_cell).split(","):
                    s = s.strip()
                    if "|" in s:
                        _, school_name = s.split("|", 1)
                        used.add(school_name.strip())
                    else:
                        used.add(s)
    return used

def load_schools():
    wb = load_workbook(SCHOOL_FILE)
    ws = wb.active
    data = {}
    used_schools = get_used_schools()

    for row in ws.iter_rows(min_row=2, values_only=True):
        _, block, cluster, school = row
        if not block or block == "block_name":
            continue
        if school in used_schools:
            continue
        data.setdefault(block, {}).setdefault(cluster, []).append(school)
    return data

@forms_bp.route("/", methods=["GET", "POST"])
def home():
    if request.method == "POST":
        aadhaar = request.files.get("aadhaar")
        photo = request.files.get("photo")
        belt_cert = request.files.get("belt_cert")
        qual_cert = request.files.get("qual_cert")

        name = request.form["name"]
        email = request.form["email"]
        mobile = request.form["mobile"]
        dob = request.form["dob"]
        address = request.form["address"]
        belt = request.form["belt"]
        qualification = request.form["qualification"]
        experience = request.form["experience"]
        block = request.form["block"]
        cluster = request.form["cluster"]
        schools = request.form.getlist("schools")

        files = {
            "aadhaar": aadhaar,
            "photo": photo,
            "belt_cert": belt_cert,
            "qual_cert": qual_cert
        }

        saved_files = {}
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)

        for key, file in files.items():
            if file and file.filename:
                filename = secure_filename(f"{name}_{key}_{file.filename}")
                path = os.path.join(UPLOAD_FOLDER, filename)
                file.save(path)
                saved_files[key] = path
            else:
                return f"{key} file is missing!"

        age = calculate_age(dob)
        trainer_id = generate_trainer_id()
        password = generate_password()
        status = "Pending"

        if not os.path.exists(TRAINER_FILE):
            wb = Workbook()
            ws = wb.active
            ws.append([
                "Full Name","Email","Mobile","DOB","Age","Address",
                "Belt","Qualification","Experience",
                "Block","Cluster","Schools",
                "Aadhaar File","Photo File","Belt Cert File","Qual Cert File",
                "Trainer ID","Password","Status"
            ])
            wb.save(TRAINER_FILE)

        wb = load_workbook(TRAINER_FILE)
        ws = wb.active
        ws.append([
            name,email,mobile,dob,age,address,
            belt,qualification,experience,
            block,cluster,", ".join(schools),
            saved_files["aadhaar"],saved_files["photo"],
            saved_files["belt_cert"],saved_files["qual_cert"],
            trainer_id,password,status
        ])
        wb.save(TRAINER_FILE)

        return redirect(url_for("forms.home", success=1))

    success = request.args.get("success") == "1"
    schools_data = load_schools()
    return render_template("index.html", schools=schools_data, success=success)


