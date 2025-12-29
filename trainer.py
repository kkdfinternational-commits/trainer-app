from flask import Blueprint, render_template_string, request
from openpyxl import load_workbook
import os

trainer_bp = Blueprint("trainer", __name__)
TRAINER_FILE = "trainers.xlsx"

def load_trainers():
    if not os.path.exists(TRAINER_FILE):
        return [], []
    wb = load_workbook(TRAINER_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return rows[0], rows[1:]

@trainer_bp.route("/trainer", methods=["GET","POST"])
def trainer_login():
    msg = None
    if request.method == "POST":
        tid = request.form["trainer_id"]
        pwd = request.form["password"]
        headers, data = load_trainers()
        for row in data:
            if row[16] == tid and row[17] == pwd:
                if row[18] == "Approved":
                    return f"Welcome {row[0]}! Trainer portal coming soon."
                else:
                    msg = "Not approved yet."
                break
        else:
            msg = "Invalid ID or password"
    return render_template_string("""
    <h2>Trainer Login</h2>
    <form method="post">
      ID:<input name="trainer_id"><br><br>
      Password:<input type="password" name="password"><br><br>
      <button>Login</button>
    </form>
    <p style="color:red;">{{msg}}</p>
    """, msg=msg)
